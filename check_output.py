import openpyxl

wb = openpyxl.load_workbook("output_test/data_catalog.xlsx")

# Find the fields sheet
sheet_name = None
for name in wb.sheetnames:
    if "tr" in name.lower():
        sheet_name = name
        break
ws = wb[sheet_name or wb.sheetnames[1]]

print(f"Sheet: {ws.title}")
print()

# Print all rows: col_name(5), type(6), length(7), key(9), values(10), mapping(13)
print(f"{'Ten_ky_thuat':<22}  {'Kieu':<12}  {'DoDai':<22}  {'Khoa':<8}  {'GiaTri':<35}  {'AnhXa'}")
print("-" * 130)
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        break
    tn   = str(row[4] or "")
    kieu = str(row[5] or "")
    dl   = str(row[6] or "")
    khoa = str(row[8] or "")
    gt   = str(row[9] or "")[:33]
    ax   = str(row[12] or "")
    print(f"{tn:<22}  {kieu:<12}  {dl:<22}  {khoa:<8}  {gt:<35}  {ax}")
