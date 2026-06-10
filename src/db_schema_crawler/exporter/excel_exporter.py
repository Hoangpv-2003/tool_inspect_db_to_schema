import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from .base import BaseExporter
from ..models.table_schema import TableSchema
from ..models.field_schema import FieldSchema

# Points per row height unit in Excel (approx. 15pt = one text line at 11pt font)
_ROW_LINE_HEIGHT_PT = 15.0
# Approximate chars that fit in one unit of column width at Calibri 11pt
_CHARS_PER_WIDTH_UNIT = 1.2


class ExcelCatalogExporter(BaseExporter):
    def __init__(self, output_path: str):
        self.output_path = output_path

    # ------------------------------------------------------------------ styles

    def _header_styles(self):
        thin = Side(border_style="thin", color="D3D3D3")
        return (
            Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
            Alignment(horizontal="center", vertical="center", wrap_text=True),
            Border(left=thin, right=thin, top=thin, bottom=thin),
        )

    def _apply_header_style(self, ws, n_cols: int) -> None:
        font, fill, align, border = self._header_styles()
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = font
            cell.fill = fill
            cell.alignment = align
            cell.border = border
        ws.row_dimensions[1].height = 30

    def _apply_column_colors(
        self,
        ws,
        auto_cols: List[int],
        manual_cols: List[int],
        n_rows: int,
    ) -> None:
        auto_fill   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        manual_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        thin   = Side(border_style="thin", color="E0E0E0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # ALL data cells get wrap_text so text never overflows visually.
        wrap_top = Alignment(vertical="top", wrap_text=True)

        all_cols = set(auto_cols) | set(manual_cols)
        for row in range(2, n_rows + 2):
            for col in all_cols:
                cell = ws.cell(row=row, column=col)
                cell.fill = auto_fill if col in auto_cols else manual_fill
                cell.border = border
                cell.alignment = wrap_top

    # ---------------------------------------------------------- column widths

    def _autofit_columns(self, ws, max_width: int = 40) -> None:
        """Set column widths based on content; cap at max_width characters."""
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells:
                val = str(cell.value or "")
                # Account for header padding
                length = len(val) + (4 if cell.row == 1 else 0)
                max_len = max(max_len, length)
            ws.column_dimensions[col_letter].width = min(max(max_len, 10), max_width)

    # ------------------------------------------------------------ row heights

    def _autofit_row_heights(self, ws, col_widths: Dict[int, float]) -> None:
        """
        Estimate and set row height for every data row so that wrapped text
        is fully visible without manual row-height adjustment in Excel.

        Formula:  lines = ceil(char_count / (col_width * CHARS_PER_WIDTH_UNIT))
                  height = max(lines) * ROW_LINE_HEIGHT_PT   (minimum 1 line)
        """
        for row in ws.iter_rows(min_row=2):
            max_lines = 1
            for cell in row:
                val = str(cell.value or "")
                if not val:
                    continue
                col_w = col_widths.get(cell.column, 15)
                chars_per_line = max(col_w * _CHARS_PER_WIDTH_UNIT, 1)
                # Each newline forces a new line regardless
                for segment in val.split("\n"):
                    lines = math.ceil(len(segment) / chars_per_line) if segment else 1
                    max_lines = max(max_lines, lines)
            ws.row_dimensions[row[0].row].height = max_lines * _ROW_LINE_HEIGHT_PT

    def _apply_all_styles(self, ws, headers, auto_cols, manual_cols, n_rows) -> None:
        """Convenience: apply header, colors, column widths, then row heights."""
        self._apply_header_style(ws, len(headers))
        self._apply_column_colors(ws, auto_cols, manual_cols, n_rows)
        self._autofit_columns(ws)

        # Collect final widths AFTER autofit so row-height calc is accurate
        col_widths = {
            col_cells[0].column: ws.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width
            for col_cells in ws.columns
        }
        self._autofit_row_heights(ws, col_widths)
        ws.freeze_panes = "A2"

    # -------------------------------------------------------------- sheets

    def _write_tables_sheet(self, ws, tables: List[TableSchema]) -> None:
        headers = [
            "STT", "Tên CSDL", "Tên bảng/dataset", "Mô tả nội dung bảng",
            "Số trường", "Số bản ghi ước tính", "Nhóm dữ liệu/thực thể", "Có dữ liệu cá nhân?",
            "Tần suất cập nhật", "Khóa định danh chính", "Người quản lý DL (Data steward)",
            "Lineage: nguồn → đích", "Ngày cập nhật gần nhất", "Thời hạn lưu trữ / hủy",
            "Giấy phép sử dụng", "Ghi chú",
        ]
        ws.append(headers)
        for table in tables:
            ws.append([
                table.stt, table.ten_csdl, table.ten_bang, table.mo_ta,
                table.so_truong, table.so_ban_ghi, table.nhom_du_lieu,
                table.co_du_lieu_ca_nhan, table.tan_suat_cap_nhat,
                table.khoa_dinh_danh, table.nguoi_quan_ly, table.lineage,
                table.ngay_cap_nhat or "", table.thoi_han_luu_tru,
                table.giay_phep, table.ghi_chu,
            ])

        # A=1, B=2, C=3, E=5, F=6, J=10, M=13  ← auto-generated
        auto_cols   = [1, 2, 3, 5, 6, 10, 13]
        # D=4, G=7, H=8, I=9, K=11, L=12, N=14, O=15, P=16 ← manual entry
        manual_cols = [4, 7, 8, 9, 11, 12, 14, 15, 16]
        self._apply_all_styles(ws, headers, auto_cols, manual_cols, len(tables))

    def _write_fields_sheet(self, ws, fields: List[FieldSchema]) -> None:
        headers = [
            "STT", "Tên CSDL", "Tên bảng/dataset", "Tên trường (nghiệp vụ)",
            "Tên trường kỹ thuật", "Kiểu dữ liệu", "Độ dài/Định dạng", "Bắt buộc?",
            "Khóa (PK/FK)", "Danh sách giá trị cho phép", "Định nghĩa nghiệp vụ",
            "Dữ liệu cá nhân?", "Ánh xạ Từ điển dùng chung", "Ghi chú",
        ]
        ws.append(headers)
        for field in fields:
            ws.append([
                field.stt, field.ten_csdl, field.ten_bang,
                field.ten_truong_nghiep_vu, field.ten_truong_ky_thuat,
                field.kieu_du_lieu, field.do_dai_dinh_dang, field.bat_buoc,
                field.khoa, field.danh_sach_gia_tri, field.dinh_nghia_nghiep_vu,
                field.du_lieu_ca_nhan, field.anh_xa, field.ghi_chu,
            ])

        # A=1, B=2, C=3, E=5, F=6, G=7, H=8, I=9, J=10, M=13 ← auto-generated
        auto_cols   = [1, 2, 3, 5, 6, 7, 8, 9, 10, 13]
        # D=4, K=11, L=12, N=14 ← manual entry
        manual_cols = [4, 11, 12, 14]
        self._apply_all_styles(ws, headers, auto_cols, manual_cols, len(fields))

    # ----------------------------------------------------------------- export

    def export(self, tables: List[TableSchema], fields: List[FieldSchema]) -> None:
        wb = openpyxl.Workbook()

        ws_tables = wb.active
        ws_tables.title = "Danh mục bảng"
        ws_tables.sheet_properties.tabColor = "1F4E79"
        self._write_tables_sheet(ws_tables, tables)

        ws_fields = wb.create_sheet(title="Danh mục trường")
        ws_fields.sheet_properties.tabColor = "375623"
        self._write_fields_sheet(ws_fields, fields)

        wb.save(self.output_path)
