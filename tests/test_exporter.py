import pytest
import openpyxl
from pathlib import Path
from db_schema_crawler.exporter.excel_exporter import ExcelCatalogExporter
from db_schema_crawler.models.table_schema import TableSchema
from db_schema_crawler.models.field_schema import FieldSchema

@pytest.fixture
def sample_data():
    tables = [
        TableSchema(
            stt=1,
            ten_csdl="DB1",
            ten_bang="users",
            mo_ta="User accounts",
            so_truong=3,
            so_ban_ghi=100,
            khoa_dinh_danh="id",
            ngay_cap_nhat="2026-06-01"
        )
    ]
    fields = [
        FieldSchema(
            stt=1,
            ten_csdl="DB1",
            ten_bang="users",
            ten_truong_ky_thuat="id",
            kieu_du_lieu="bigint",
            do_dai_dinh_dang="20 chữ số",
            bat_buoc="Có",
            khoa="PK"
        ),
        FieldSchema(
            stt=2,
            ten_csdl="DB1",
            ten_bang="users",
            ten_truong_ky_thuat="email",
            kieu_du_lieu="varchar",
            do_dai_dinh_dang="100 ký tự",
            bat_buoc="Có",
            khoa=""
        )
    ]
    return tables, fields

def test_export_creates_single_file(tmp_path, sample_data):
    tables, fields = sample_data
    file_path = tmp_path / "data_catalog.xlsx"
    
    exporter = ExcelCatalogExporter(str(file_path))
    exporter.export(tables, fields)
    
    assert file_path.exists()
    assert len(list(tmp_path.iterdir())) == 1

def test_export_has_two_sheets(tmp_path, sample_data):
    tables, fields = sample_data
    file_path = tmp_path / "data_catalog.xlsx"
    
    exporter = ExcelCatalogExporter(str(file_path))
    exporter.export(tables, fields)
    
    wb = openpyxl.load_workbook(str(file_path))
    assert len(wb.sheetnames) == 2
    assert "Danh mục bảng" in wb.sheetnames
    assert "Danh mục trường" in wb.sheetnames

def test_tables_sheet_header_16_cols(tmp_path, sample_data):
    tables, fields = sample_data
    file_path = tmp_path / "data_catalog.xlsx"
    
    exporter = ExcelCatalogExporter(str(file_path))
    exporter.export(tables, fields)
    
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb["Danh mục bảng"]
    assert ws.max_column == 16
    
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "STT"
    assert headers[2] == "Tên bảng/dataset"
    assert headers[9] == "Khóa định danh chính"

def test_fields_sheet_header_14_cols(tmp_path, sample_data):
    tables, fields = sample_data
    file_path = tmp_path / "data_catalog.xlsx"
    
    exporter = ExcelCatalogExporter(str(file_path))
    exporter.export(tables, fields)
    
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb["Danh mục trường"]
    assert ws.max_column == 14
    
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "STT"
    assert headers[4] == "Tên trường kỹ thuật"
    assert headers[8] == "Khóa (PK/FK)"

def test_tables_sheet_row_count(tmp_path, sample_data):
    tables, fields = sample_data
    file_path = tmp_path / "data_catalog.xlsx"
    
    exporter = ExcelCatalogExporter(str(file_path))
    exporter.export(tables, fields)
    
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb["Danh mục bảng"]
    # Row 1 is header, Row 2 is data
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=3).value == "users"

def test_fields_sheet_row_count(tmp_path, sample_data):
    tables, fields = sample_data
    file_path = tmp_path / "data_catalog.xlsx"
    
    exporter = ExcelCatalogExporter(str(file_path))
    exporter.export(tables, fields)
    
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb["Danh mục trường"]
    # Row 1 header, Row 2 & 3 data
    assert ws.max_row == 3
    assert ws.cell(row=2, column=5).value == "id"
    assert ws.cell(row=3, column=5).value == "email"

def test_tables_auto_columns_filled(tmp_path, sample_data):
    tables, fields = sample_data
    file_path = tmp_path / "data_catalog.xlsx"
    
    exporter = ExcelCatalogExporter(str(file_path))
    exporter.export(tables, fields)
    
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb["Danh mục bảng"]
    # Auto column E (column 5) - number of fields
    assert ws.cell(row=2, column=5).value == 3
    # Auto column F (column 6) - record count
    assert ws.cell(row=2, column=6).value == 100

def test_tables_manual_columns_empty(tmp_path, sample_data):
    tables, fields = sample_data
    file_path = tmp_path / "data_catalog.xlsx"
    
    exporter = ExcelCatalogExporter(str(file_path))
    exporter.export(tables, fields)
    
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb["Danh mục bảng"]
    # Column G (column 7) - nhom_du_lieu (default empty)
    assert ws.cell(row=2, column=7).value in ("", None)

def test_header_background_color(tmp_path, sample_data):
    tables, fields = sample_data
    file_path = tmp_path / "data_catalog.xlsx"
    
    exporter = ExcelCatalogExporter(str(file_path))
    exporter.export(tables, fields)
    
    wb = openpyxl.load_workbook(str(file_path))
    ws = wb["Danh mục bảng"]
    header_fill = ws.cell(row=1, column=1).fill
    # openpyxl returns fill start_color.rgb
    assert header_fill.start_color.rgb == "001F4E79" or header_fill.start_color.rgb == "1F4E79"

def test_freeze_panes_row2(tmp_path, sample_data):
    tables, fields = sample_data
    file_path = tmp_path / "data_catalog.xlsx"
    
    exporter = ExcelCatalogExporter(str(file_path))
    exporter.export(tables, fields)
    
    wb = openpyxl.load_workbook(str(file_path))
    assert wb["Danh mục bảng"].freeze_panes == "A2"
    assert wb["Danh mục trường"].freeze_panes == "A2"
